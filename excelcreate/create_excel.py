from openpyxl import Workbook, load_workbook 

# write data to execl sheet
def write_excel(filename, data):
    wb =Workbook()  
    sheet = wb.active  

    for row in data:  
        sheet.append(row)

    wb.save(filename) 
    print(f"Data saved in '{filename}' successfully.")
    
x=write_excel("myecxcel.xlsx", [["Name", "Age"], ["nour", 25], ["ali", 30]])  
print(x)
#######################################################################
#read data from my sheet 
def read_excel(filename):
    wb =load_workbook(filename) 
    sheet = wb.active  

    for row in sheet.iter_rows(values_only=True):  
        print(row)
y=read_excel("myecxcel.xlsx")  
print(y)
#############################################################################################3
# append data to my excel sheet
def append_to_excel(filename, data):
    wb = load_workbook(filename)  
    sheet = wb.active  

    for row in data:  
        sheet.append(row)

    wb.save(filename)  

z=append_to_excel("myecxcel.xlsx", [["noha", 22], ["amr", 26]])  
print(z)
c=read_excel("myecxcel.xlsx")  
print(c)